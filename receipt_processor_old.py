import cv2
import pytesseract
from PIL import Image
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# Set the Tesseract command path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def preprocess_image(image_path):
    try:
        img = cv2.imread(image_path)
        if img is None:
            raise ValueError("Unable to read the image. The file may be corrupted or in an unsupported format.")
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        return thresh
    except Exception as e:
        print(f"Error in image preprocessing: {str(e)}")
        return None

def extract_text(preprocessed_image):
    try:
        return pytesseract.image_to_string(Image.fromarray(preprocessed_image))
    except Exception as e:
        print(f"Error in text extraction: {str(e)}")
        return None

def categorize_spending(total, date):
    try:
        total = float(total)
        date = datetime.strptime(date, "%d/%m/%Y")
        
        if total < 20:
            category = "Low-cost"
        elif 20 <= total < 100:
            category = "Medium-cost"
        else:
            category = "High-cost"
        
        if date.month in [12, 1, 2]:
            category += " Winter"
        elif date.month in [3, 4, 5]:
            category += " Spring"
        elif date.month in [6, 7, 8]:
            category += " Summer"
        else:
            category += " Fall"
        
        return category
    except Exception as e:
        print(f"Error in categorization: {str(e)}")
        return "Uncategorized"

def parse_receipt(text):
    try:
        text_lower = text.lower()
        
        date_match = re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', text_lower)
        date = date_match.group(0) if date_match else "Date not found"
        
        total_match = re.search(r'total:?\s*\$?(\d+\.?\d{2})', text_lower)
        total = total_match.group(1) if total_match else "0.00"
        
        category = categorize_spending(total, date)
        
        return {
            'Date': date,
            'Total': total,
            'Category': category
        }
    except Exception as e:
        print(f"Error in parsing receipt: {str(e)}")
        return None

def user_confirm_data(data):
    print("\nExtracted Data:")
    for key, value in data.items():
        print(f"{key}: {value}")
    
    while True:
        confirm = input("\nIs this information correct? (yes/no): ").lower()
        if confirm in ['yes', 'y']:
            return data
        elif confirm in ['no', 'n']:
            print("\nPlease enter the correct information:")
            data['Date'] = input("Date (DD/MM/YYYY): ")
            data['Total'] = input("Total amount: ")
            data['Category'] = categorize_spending(data['Total'], data['Date'])
            return data
        else:
            print("Invalid input. Please enter 'yes' or 'no'.")

def save_to_excel(data, output_file='receipts.xlsx'):
    try:
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Date', 'Total', 'Category'])

        ws.append([data['Date'], data['Total'], data['Category']])
        wb.save(output_file)
        print(f"Data saved to {output_file}")
        print(f"Full path of saved file: {os.path.abspath(output_file)}")
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")

def process_receipt(image_path):
    preprocessed = preprocess_image(image_path)
    if preprocessed is None:
        return None

    text = extract_text(preprocessed)
    

    if text is None:
        return None

    print(text);
    print('---end----');
    return parse_receipt(text)

# Main execution
def main():
    while True:
        image_path = input("Enter the path to the receipt image (or 'quit' to exit): ")
        if image_path.lower() == 'quit':
            break
        receipt_data = process_receipt(image_path)
        # receipt_data = process_receipt('C:\Users\User\Desktop\receipts\receipt1.jpeg');
        if receipt_data:
            confirmed_data = user_confirm_data(receipt_data)
            save_to_excel(confirmed_data)
        else:
            print("Failed to process the receipt. Please try again with a different image.")

if __name__ == "__main__":
    main()